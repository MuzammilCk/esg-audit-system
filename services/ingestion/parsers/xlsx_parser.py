from __future__ import annotations

import base64
from io import BytesIO
from typing import Any, Dict, List

from services.ingestion.models import NormalizedImage, NormalizedTable


def _stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_xlsx(raw_xlsx: bytes) -> Dict[str, Any]:
    """Parse an XLSX into text, tables, images, and structure."""

    from openpyxl import load_workbook

    wb = load_workbook(filename=BytesIO(raw_xlsx), data_only=True)

    tables: List[NormalizedTable] = []
    images: List[NormalizedImage] = []
    sections: List[Dict[str, Any]] = []
    text_parts: List[str] = []

    for ws in wb.worksheets:
        # Tables
        data: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            row_data = [_stringify_cell(v) for v in row]
            if any(c != "" for c in row_data):
                data.append(row_data)

        if data:
            tables.append(NormalizedTable(page=1, data=data))

            preview = None
            try:
                from tabulate import tabulate

                if len(data) >= 2:
                    preview = tabulate(data[:10], headers="firstrow", tablefmt="github")
                else:
                    preview = tabulate(data, tablefmt="github")
            except Exception:
                preview = None

            sections.append(
                {
                    "kind": "sheet",
                    "title": ws.title,
                    "page": 1,
                    "table_preview": preview,
                }
            )

            text_parts.append(f"[Sheet: {ws.title}]")
            if preview:
                text_parts.append(preview)

        # Images (best-effort; page=1)
        for img in getattr(ws, "_images", []) or []:
            img_bytes = b""
            try:
                img_bytes = img._data()  # type: ignore[attr-defined]
            except Exception:
                # Fallback: if PIL image is attached
                try:
                    pil_img = getattr(img, "image", None)
                    if pil_img is not None:
                        buf = BytesIO()
                        pil_img.save(buf, format="PNG")
                        img_bytes = buf.getvalue()
                except Exception:
                    img_bytes = b""

            if img_bytes:
                images.append(
                    NormalizedImage(
                        page=1,
                        base64=base64.b64encode(img_bytes).decode("ascii"),
                        caption=ws.title,
                    )
                )

    text_content = "\n".join(text_parts).strip()

    return {
        "text_content": text_content,
        "tables": tables,
        "images": images,
        "pages": 1,
        "sections": sections,
    }
